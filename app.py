from flask import Flask, render_template, request, redirect, url_for, session, send_file
from fpdf import FPDF
import openpyxl, io, datetime

app = Flask(__name__)
app.secret_key = 'attendance2024secret'

ADMIN = {'username': 'admin', 'password': 'admin123'}

PROFILE = {
    'name': 'Prof. Harshali Vihire',
    'email': 'admin@company.com',
    'role': 'HR Administrator',
    'department': 'Human Resources',
    'phone': '+91 98765 43210'
}

employees = []
attendance = []

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

@app.route('/', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        if request.form['username'] == ADMIN['username'] and request.form['password'] == ADMIN['password']:
            session['logged_in'] = True
            return redirect(url_for('dashboard'))
        error = 'Invalid username or password'
    return render_template('login.html', error=error)

@app.route('/dashboard')
@login_required
def dashboard():
    total_emp = len(employees)
    today = datetime.date.today().strftime('%Y-%m-%d')
    today_records = [a for a in attendance if a['date'] == today]
    present = sum(1 for a in today_records if a['status'] == 'Present')
    absent = sum(1 for a in today_records if a['status'] == 'Absent')
    return render_template('dashboard.html',
        employees=employees, attendance=attendance,
        total_emp=total_emp, present=present, absent=absent,
        today=today)

@app.route('/add_employee', methods=['POST'])
@login_required
def add_employee():
    name = request.form['name'].strip()
    dept = request.form['department'].strip()
    emp_id = len(employees) + 1
    employees.append({'id': emp_id, 'name': name, 'department': dept})
    return redirect(url_for('dashboard'))

@app.route('/delete_employee/<int:eid>')
@login_required
def delete_employee(eid):
    global employees
    employees = [e for e in employees if e['id'] != eid]
    return redirect(url_for('dashboard'))

@app.route('/mark_attendance', methods=['POST'])
@login_required
def mark_attendance():
    emp_id = int(request.form['emp_id'])
    status = request.form['status']
    date = datetime.date.today().strftime('%Y-%m-%d')
    emp = next((e for e in employees if e['id'] == emp_id), None)
    if emp:
        existing = next((a for a in attendance if a['emp_id'] == emp_id and a['date'] == date), None)
        if existing:
            existing['status'] = status
        else:
            attendance.append({
                'id': len(attendance) + 1,
                'emp_id': emp_id,
                'name': emp['name'],
                'department': emp['department'],
                'date': date,
                'status': status
            })
    return redirect(url_for('dashboard'))

@app.route('/profile')
@login_required
def profile():
    return render_template('profile.html', profile=PROFILE)

@app.route('/export/pdf')
@login_required
def export_pdf():
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 16)
    pdf.cell(0, 12, 'Employee Attendance Report', ln=True, align='C')
    pdf.set_font('Helvetica', 'B', 11)
    pdf.set_fill_color(26, 115, 232)
    pdf.set_text_color(255, 255, 255)
    for col, w in [('ID',15),('Name',50),('Department',45),('Date',35),('Status',35)]:
        pdf.cell(w, 10, col, border=1, fill=True)
    pdf.ln()
    pdf.set_font('Helvetica', '', 10)
    pdf.set_text_color(0, 0, 0)
    for a in attendance:
        pdf.set_fill_color(240,240,240) if a['id'] % 2 == 0 else pdf.set_fill_color(255,255,255)
        pdf.cell(15, 9, str(a['id']), border=1, fill=True)
        pdf.cell(50, 9, a['name'], border=1, fill=True)
        pdf.cell(45, 9, a['department'], border=1, fill=True)
        pdf.cell(35, 9, a['date'], border=1, fill=True)
        pdf.cell(35, 9, a['status'], border=1, fill=True)
        pdf.ln()
    buf = io.BytesIO(pdf.output())
    return send_file(buf, mimetype='application/pdf', download_name='attendance.pdf', as_attachment=True)

@app.route('/export/excel')
@login_required
def export_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance'
    ws.append(['ID', 'Name', 'Department', 'Date', 'Status'])
    for a in attendance:
        ws.append([a['id'], a['name'], a['department'], a['date'], a['status']])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='attendance.xlsx', as_attachment=True)

@app.route('/health')
def health():
    return {'status': 'UP', 'employees': len(employees), 'records': len(attendance)}, 200

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)